
# -*- coding: utf-8 -*-
#%%[markdown]
# # Simplificando pandas, lectura y escritura de archivos III
# ## Lectura y manipulación de archivos excel.
#%%[markdown]
# ### Importamos las librerias a utilizar
#%%
import os
from pathlib import Path
import glob
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import xlwings as xw
#%%[markdown]
# # Módulo pathlib
#%%
# directorio actual
dir_prin = Path.cwd()
# directorio con los datos
dir_entra = dir_prin/'entradas'
# directorio con los datos intermedios
dir_inter = dir_prin /'intermedios'
# directorio con los datos finales
dir_salid = dir_prin / 'salidas'
#%%[markdown]
# ## Lectura de una sola pestaña
# ### Pestaña por defecto
#%%
filename = dir_entra / 'datos_clima.xlsx'
datos = pd.read_excel(filename)
datos.head()
#%%[markdown]
# ### Escogiendo la pestaña
#%%
datos = pd.read_excel(filename, sheet_name='febrero') 
datos.head()
#%%[markdown]
# ## Lectura de varias pestañas de un mismo archivo
# ### Primero conozcamos las pestañas del archivo
#%%[markdown]
# Ocupamos openpyxl
#
# Creamos el objeto wb con la información del archivo
#%%
wb = load_workbook(filename)
#%%[markdown]
# Extraemos el nombre de las pestañas
#%%
meses = wb.sheetnames
meses
#%%[markdown]
# ### Dataframe vacío
#%%
datos = pd.DataFrame({'time':[],'precipIntensity':[],
                    'temperature':[],'apparentTemperature':[],
                    'dewPoint':[],'humidity':[],'pressure':[],
                    'windSpeed':[],'cloudCover':[]})
columnas = ['time','precipIntensity','temperature','apparentTemperature',
            'dewPoint','humidity','pressure','windSpeed','cloudCover']
#%%
for mes in meses:
    datos_temp = pd.read_excel(filename, sheet_name=mes)
    # Seleccionamos las columnas de interes
    datos_temp = datos_temp[columnas]
    # Unimos Dataframnes
    datos = pd.concat([datos, datos_temp], join='inner', axis=0)
    datos = datos[columnas]
datos = datos.set_index('time')
#%%[markdown]
# ### Revisemos los datos
#%%
datos.head()
#%%
datos.tail()
#%%
datos.shape
#%%[markdown]
# ## Selección de una columna
# Veremos tres formas de seleccionar una columna:
# ### Usando corchetes y nombre entre comillas
#%%
datos['precipIntensity'].describe()
#%%[markdown]
# ### Usando notación de punto
#%%
datos.precipIntensity.describe()
#%%[markdown]
# ### Usando "iloc" y el número de la columna
#%%
datos.iloc[:,0].describe()
#%%[markdown]
# ### Otras funciones estadísticas
#%%
[datos['precipIntensity'].sum(), # Suma del valor de las columnas
 datos['precipIntensity'].nunique(), # Número de valores únicos
 datos['precipIntensity'].isnull().sum(), # Total de valores nulos
 datos['precipIntensity'].notnull().sum() # Total de valores no nulos
 ] 
#%%
datos['precipIntensity'].value_counts()
#%%
type(datos['precipIntensity'])
#%%[markdown]
 # ## Selección de múltiples columnas
# Veremos dos formas de seleccionar una columna:
# ### Usando corchetes y nombre entre comillas
#%%
datos[['precipIntensity','temperature','apparentTemperature']].describe()
#%%[markdown]
# ### Usando "iloc" y el número de las columnas
#%%
datos.iloc[:,0:3].describe()
#%%[markdown]
# ## Selección de filas
# Veremos tres formas de hacerlo
# ### Usando iloc y el número de la fila
#%%
datos.iloc[0]
#%%
datos.iloc[-1]
#%%
type(datos.iloc[-1])
#%%[markdown]
# ### Usando el índice de la fila
#%%
datos['2018-01-01 06:00:00']
#%%
type(datos['2018-01-01 06:00:00'])
#%%[markdown]
# ## Seleccionando múltiples filas
#%%
datos.iloc[0:3]
#%%[markdown]
# ## Seleccionando una porción del dataframe
# Veremos dos formas:
# ### Usando iloc
#%%
datos.iloc[[0,3,6,-1], [0,5,-3]] 
#%%[markdown]
# ### Usando corchetes e iloc
#%%
datos[['precipIntensity','pressure','pressure']].iloc[[0,3,6,-1]]
#%%[markdown]
# ## Cambiando el nombre de las columnas
# Los nombres actuales de las columnas son:
#%%
datos.columns
#%%[markdown]
# ### Poniendo en minúsculas todos los nombres de columnas
#%%
datos = datos.rename(columns=str.lower) 
datos.columns

# datos.rename(columns=str.lower, inplace = True)
#%%
datos = datos.rename(
    columns = {
        'precipintensity': 'precip_intensity',
        'apparenttemperature': 'apparent_temperature',
        'dewpoint': 'dew_point',
        'windspeed': 'wind_speed',
        'cloudcover': 'cloud_cover'
    }
)
datos.columns
#%%[markdown]
# ## Escritura de archivos excel
# ### Sin formato
#%%
reporte = dir_inter / 'clima2018.xlsx'
writer = pd.ExcelWriter(reporte, engine='xlsxwriter')
datos.to_excel(writer,'Datos', index=True)
writer.save()
#%%[markdown]
# ### Con formato
#%%
reporte = dir_inter / 'clima2018_format.xlsx'
writer = pd.ExcelWriter(reporte, engine='xlsxwriter')
workbook  = writer.book
datos.to_excel(writer,'Datos', index=True)
worksheet = writer.sheets['Datos']
#Creamos los formatos condicionales
format1 = workbook.add_format({'bg_color':   '#FFC7CE',
                            'font_color': '#9C0006'})
format2 = workbook.add_format({'bg_color':   '#C6EFCE',
                                       'font_color': '#006100'})
#Aplicamos los formatos
rango1 = 'B1:I' + str(len(datos) + 1)
worksheet.conditional_format(rango1, {'type': 'blanks',
                                        'format':   format1})   
rango2 = 'B1:I' + str(len(datos) + 1)   
worksheet.conditional_format(rango2, {'type':     'cell',
                                       'criteria': 'between',
                                       'minimum':  25,
                                       'maximum':  28,
                                       'format':   format2}) 
writer.save()
#Autoajustar el ancho de la columna
#Librería xlwings
reporte2 = str(reporte)
wb = xw.Book(reporte2)
sht = wb.sheets['Datos']
sht.autofit('c')
wb.save(reporte2)
#%%[markdown]
# #SPAM
# Muchas gracias, no te olvides de darle like, suscribirte y activar las notificaciones